"""
AgroPrix — Export de rapports institutionnels (PDF + Excel).

Génère des rapports structurés pour coopératives, ONG, bailleurs.
PDF : ReportLab avec branding AgroPrix/33Lab.
Excel : openpyxl avec plusieurs onglets (prix, marchés, statistiques, couverture EUDR).
"""

from datetime import datetime, timezone
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import text

from app.auth import get_current_user
from app.database import get_engine, sql_year_month
from app.config import UEMOA_COUNTRIES

router = APIRouter(prefix="", tags=["rapports"])

# AgroPrix brand colors (hex)
GREEN_HEX = "#2d8a4e"
NAVY_HEX  = "#1a3a5c"
GOLD_HEX  = "#f4a261"


# ---------------------------------------------------------------------------
# Helpers — data fetching
# ---------------------------------------------------------------------------

def _fetch_prices(engine, country: str, commodity: str, start_date: Optional[str], end_date: Optional[str]):
    conditions = ["country = :country", "commodity LIKE :commodity"]
    params = {"country": country, "commodity": f"%{commodity}%"}
    if start_date:
        conditions.append("date >= :start_date")
        params["start_date"] = start_date
    if end_date:
        conditions.append("date <= :end_date")
        params["end_date"] = end_date
    where = " AND ".join(conditions)
    q = text(f"""
        SELECT date, market, commodity, price, currency, unit, source, latitude, longitude
        FROM prices WHERE {where} ORDER BY date DESC LIMIT 500
    """)
    with engine.connect() as conn:
        return [dict(r._mapping) for r in conn.execute(q, params)]


def _fetch_monthly(engine, country: str, commodity: str, start_date: Optional[str], end_date: Optional[str]):
    conditions = ["country = :country", "commodity LIKE :commodity"]
    params = {"country": country, "commodity": f"%{commodity}%"}
    if start_date:
        conditions.append("date >= :start_date")
        params["start_date"] = start_date
    if end_date:
        conditions.append("date <= :end_date")
        params["end_date"] = end_date
    where = " AND ".join(conditions)
    q = text(f"""
        SELECT {sql_year_month('date')} as mois,
               ROUND(AVG(price), 1) as prix_moyen,
               MIN(price) as prix_min,
               MAX(price) as prix_max,
               COUNT(DISTINCT market) as nb_marches,
               COUNT(*) as nb_observations
        FROM prices WHERE {where}
        GROUP BY {sql_year_month('date')}
        ORDER BY mois
    """)
    with engine.connect() as conn:
        return [dict(r._mapping) for r in conn.execute(q, params)]


def _fetch_compare(engine, commodity: str):
    q = text(f"""
        SELECT country, ROUND(AVG(price), 1) as prix_moyen,
               COUNT(DISTINCT market) as nb_marches, MAX(date) as derniere_date
        FROM prices WHERE commodity LIKE :commodity
        GROUP BY country ORDER BY prix_moyen DESC
    """)
    with engine.connect() as conn:
        return [dict(r._mapping) for r in conn.execute(q, {"commodity": f"%{commodity}%"})]


# ---------------------------------------------------------------------------
# PDF Report
# ---------------------------------------------------------------------------

@router.get(
    "/pdf",
    summary="Rapport institutionnel PDF",
    description=(
        "Génère un rapport PDF complet : prix historiques, moyennes mensuelles, "
        "comparaison régionale UEMOA, recommandations. "
        "Format A4 avec branding AgroPrix/33Lab. "
        "Adapté aux rapports trimestriels pour coopératives, ONG et bailleurs."
    ),
)
async def export_pdf(
    country: str = Query(..., description="Pays (ex: benin, cote_divoire)"),
    commodity: str = Query(..., description="Culture (ex: maize, cacao, cajou)"),
    start_date: Optional[str] = Query(None, description="Début YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="Fin YYYY-MM-DD"),
    current_user: dict = Depends(get_current_user),
):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm, cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    GREEN = colors.HexColor(GREEN_HEX)
    NAVY  = colors.HexColor(NAVY_HEX)
    GOLD  = colors.HexColor(GOLD_HEX)
    LIGHT = colors.HexColor("#f0f7f4")

    engine = get_engine()
    prices_data  = _fetch_prices(engine, country, commodity, start_date, end_date)
    monthly_data = _fetch_monthly(engine, country, commodity, start_date, end_date)
    compare_data = _fetch_compare(engine, commodity)

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            topMargin=2*cm, bottomMargin=2*cm,
                            leftMargin=2*cm, rightMargin=2*cm)

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle("APTitle",   fontName="Helvetica-Bold", fontSize=22,
                              textColor=GREEN, spaceAfter=4))
    styles.add(ParagraphStyle("APSub",     fontName="Helvetica",      fontSize=11,
                              textColor=NAVY, spaceAfter=12))
    styles.add(ParagraphStyle("APSection", fontName="Helvetica-Bold", fontSize=13,
                              textColor=NAVY, spaceBefore=14, spaceAfter=6))
    styles.add(ParagraphStyle("APBody",    fontName="Helvetica",      fontSize=9,
                              textColor=colors.black, spaceAfter=4, leading=13))
    styles.add(ParagraphStyle("APFooter",  fontName="Helvetica",      fontSize=7,
                              textColor=colors.grey, alignment=TA_CENTER))
    styles.add(ParagraphStyle("APBadge",   fontName="Helvetica-Bold", fontSize=9,
                              textColor=colors.white, alignment=TA_CENTER))

    country_names = {
        "benin": "Bénin", "burkina_faso": "Burkina Faso",
        "cote_divoire": "Côte d'Ivoire", "guinee_bissau": "Guinée-Bissau",
        "mali": "Mali", "niger": "Niger", "senegal": "Sénégal", "togo": "Togo",
    }
    cn = country_names.get(country, country.replace("_", " ").title())
    iso3 = UEMOA_COUNTRIES.get(country, {}).get("iso3", "")
    _today = "aujourd'hui"
    period = f"{start_date or 'début'} → {end_date or _today}"
    now_str = datetime.now().strftime("%d/%m/%Y %H:%M")

    elems = []

    # ── Header ──────────────────────────────────────────────────────────────
    elems.append(Paragraph("AgroPrix — Rapport Institutionnel", styles["APTitle"]))
    elems.append(Paragraph(
        f"<b>Pays :</b> {cn} ({iso3}) &nbsp;|&nbsp; "
        f"<b>Filière :</b> {commodity.title()} &nbsp;|&nbsp; "
        f"<b>Période :</b> {period}",
        styles["APSub"],
    ))
    elems.append(Paragraph(
        f"Généré le {now_str} par <b>{current_user.get('name', 'AgroPrix')}</b> — "
        "Source : WFP DataBridges + terrain AgroPrix",
        styles["APBody"],
    ))
    elems.append(HRFlowable(width="100%", thickness=2, color=GREEN, spaceAfter=10))

    # ── Stats globales ───────────────────────────────────────────────────────
    if prices_data:
        all_p = [r["price"] for r in prices_data if r.get("price")]
        avg_p = sum(all_p) / len(all_p) if all_p else 0
        nb_m  = len({r["market"] for r in prices_data})

        stat_data = [
            ["Observations", "Marchés", "Prix moyen", "Prix min", "Prix max"],
            [str(len(all_p)), str(nb_m),
             f"{avg_p:,.0f} XOF", f"{min(all_p):,.0f} XOF", f"{max(all_p):,.0f} XOF"],
        ]
        st = Table(stat_data, colWidths=[70, 70, 90, 80, 80])
        st.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0), GREEN),
            ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
            ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME",    (0, 1), (-1, 1), "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, -1), 9),
            ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
            ("BACKGROUND",  (0, 1), (-1, 1), LIGHT),
            ("GRID",        (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("TOPPADDING",  (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        elems.append(st)
        elems.append(Spacer(1, 6*mm))

    # ── Moyennes mensuelles ──────────────────────────────────────────────────
    if monthly_data:
        elems.append(Paragraph("Évolution Mensuelle des Prix", styles["APSection"]))
        t_data = [["Mois", "Prix moyen (XOF)", "Min", "Max", "Marchés", "Observations"]]
        for r in monthly_data:
            t_data.append([
                r["mois"], f"{r['prix_moyen']:,.0f}",
                f"{r['prix_min']:,.0f}", f"{r['prix_max']:,.0f}",
                str(r["nb_marches"]), str(r["nb_observations"]),
            ])
        t = Table(t_data, colWidths=[60, 90, 65, 65, 55, 75])
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0), NAVY),
            ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
            ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, -1), 8),
            ("ALIGN",         (1, 0), (-1, -1), "RIGHT"),
            ("GRID",          (0, 0), (-1, -1), 0.3, colors.lightgrey),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, LIGHT]),
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elems.append(t)
        elems.append(Spacer(1, 6*mm))

    # ── Comparaison régionale ────────────────────────────────────────────────
    if compare_data:
        elems.append(Paragraph("Comparaison Régionale UEMOA", styles["APSection"]))
        c_data = [["Pays", "Prix moyen (XOF)", "Marchés", "Dernière mise à jour"]]
        for r in compare_data:
            c_data.append([
                country_names.get(r["country"], r["country"].title()),
                f"{r['prix_moyen']:,.0f}", str(r["nb_marches"]), r["derniere_date"] or "—",
            ])
        ct = Table(c_data, colWidths=[120, 110, 70, 110])
        ct.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0), GOLD),
            ("TEXTCOLOR",     (0, 0), (-1, 0), NAVY),
            ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, -1), 8),
            ("ALIGN",         (1, 0), (1, -1), "RIGHT"),
            ("ALIGN",         (2, 0), (2, -1), "CENTER"),
            ("GRID",          (0, 0), (-1, -1), 0.3, colors.lightgrey),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, LIGHT]),
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elems.append(ct)
        elems.append(Spacer(1, 6*mm))

    # ── Prix récents (top 20) ────────────────────────────────────────────────
    if prices_data:
        elems.append(Paragraph("Derniers Prix Observés (20 plus récents)", styles["APSection"]))
        p_data = [["Date", "Marché", "Prix (XOF)", "Unité", "Source"]]
        for r in prices_data[:20]:
            p_data.append([
                r.get("date", "—"), r.get("market", "—"),
                f"{r.get('price', 0):,.0f}", r.get("unit", "KG"), r.get("source", "—"),
            ])
        pt = Table(p_data, colWidths=[65, 160, 80, 45, 60])
        pt.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0), GREEN),
            ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
            ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, -1), 8),
            ("ALIGN",         (2, 0), (2, -1), "RIGHT"),
            ("GRID",          (0, 0), (-1, -1), 0.3, colors.lightgrey),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, LIGHT]),
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elems.append(pt)

    # ── Footer ───────────────────────────────────────────────────────────────
    elems.append(Spacer(1, 12*mm))
    elems.append(HRFlowable(width="100%", thickness=1, color=colors.lightgrey))
    elems.append(Spacer(1, 3*mm))
    elems.append(Paragraph(
        "AgroPrix by 33Lab · Cotonou, Bénin · agroprix.app · api@agroprix.app",
        styles["APFooter"],
    ))
    elems.append(Paragraph(
        "Données : WFP DataBridges (vam-data-bridges) + contributions terrain. "
        "Compatible ECOAGRIS/ECOWAS. Rapport généré automatiquement.",
        styles["APFooter"],
    ))

    doc.build(elems)
    buf.seek(0)

    fname = f"agroprix_{country}_{commodity}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ---------------------------------------------------------------------------
# Excel Report
# ---------------------------------------------------------------------------

@router.get(
    "/excel",
    summary="Rapport institutionnel Excel (.xlsx)",
    description=(
        "Génère un fichier Excel multi-onglets : "
        "Prix bruts · Moyennes mensuelles · Comparaison UEMOA · Couverture données. "
        "Idéal pour l'analyse par les équipes terrain des bailleurs (Enabel, USAID, GIZ, AFD, IFAD, BAD)."
    ),
)
async def export_excel(
    country: str = Query(..., description="Pays (ex: benin)"),
    commodity: str = Query(..., description="Culture (ex: maize, cacao, cajou)"),
    start_date: Optional[str] = Query(None, description="Début YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="Fin YYYY-MM-DD"),
    current_user: dict = Depends(get_current_user),
):
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers,
    )
    from openpyxl.utils import get_column_letter

    engine = get_engine()
    prices_data  = _fetch_prices(engine, country, commodity, start_date, end_date)
    monthly_data = _fetch_monthly(engine, country, commodity, start_date, end_date)
    compare_data = _fetch_compare(engine, commodity)
    _today = "aujourd'hui"

    wb = openpyxl.Workbook()

    # Styles
    hdr_fill  = PatternFill("solid", fgColor="2d8a4e")
    hdr2_fill = PatternFill("solid", fgColor="1a3a5c")
    gold_fill = PatternFill("solid", fgColor="f4a261")
    alt_fill  = PatternFill("solid", fgColor="f0f7f4")
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    hdr2_font = Font(bold=True, color="FFFFFF", size=10)
    body_font = Font(size=9)
    bold_font = Font(bold=True, size=9)
    thin      = Side(style="thin", color="CCCCCC")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    center    = Alignment(horizontal="center", vertical="center")

    def style_header(ws, row_num, fill, font):
        for cell in ws[row_num]:
            cell.fill = fill
            cell.font = font
            cell.alignment = center
            cell.border = border

    def auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    country_names = {
        "benin": "Bénin", "burkina_faso": "Burkina Faso",
        "cote_divoire": "Côte d'Ivoire", "guinee_bissau": "Guinée-Bissau",
        "mali": "Mali", "niger": "Niger", "senegal": "Sénégal", "togo": "Togo",
    }

    # ── Onglet 1 : Prix bruts ─────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Prix bruts"
    headers1 = ["Date", "Marché", "Produit", "Prix (XOF)", "Unité", "Source", "Latitude", "Longitude"]
    ws1.append(headers1)
    style_header(ws1, 1, hdr_fill, hdr_font)
    for r in prices_data:
        ws1.append([
            r.get("date"), r.get("market"), r.get("commodity"),
            r.get("price"), r.get("unit", "KG"), r.get("source"),
            r.get("latitude"), r.get("longitude"),
        ])
    for i, row in enumerate(ws1.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.font = body_font
            cell.border = border
            if i % 2 == 0:
                cell.fill = alt_fill
    auto_width(ws1)

    # ── Onglet 2 : Moyennes mensuelles ────────────────────────────────────
    ws2 = wb.create_sheet("Moyennes mensuelles")
    headers2 = ["Mois", "Prix moyen (XOF)", "Prix min", "Prix max", "Nb marchés", "Nb observations"]
    ws2.append(headers2)
    style_header(ws2, 1, hdr2_fill, hdr2_font)
    for r in monthly_data:
        ws2.append([
            r["mois"], r["prix_moyen"], r["prix_min"],
            r["prix_max"], r["nb_marches"], r["nb_observations"],
        ])
    for i, row in enumerate(ws2.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.font = body_font
            cell.border = border
            if i % 2 == 0:
                cell.fill = alt_fill
    auto_width(ws2)

    # ── Onglet 3 : Comparaison UEMOA ─────────────────────────────────────
    ws3 = wb.create_sheet("Comparaison UEMOA")
    headers3 = ["Pays", "ISO3", "Prix moyen (XOF)", "Nb marchés", "Dernière date"]
    ws3.append(headers3)
    style_header(ws3, 1, gold_fill, Font(bold=True, color="1a3a5c", size=10))
    for r in compare_data:
        ws3.append([
            country_names.get(r["country"], r["country"].title()),
            UEMOA_COUNTRIES.get(r["country"], {}).get("iso3", ""),
            r["prix_moyen"], r["nb_marches"], r["derniere_date"],
        ])
    for i, row in enumerate(ws3.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.font = body_font
            cell.border = border
            if i % 2 == 0:
                cell.fill = alt_fill
    auto_width(ws3)

    # ── Onglet 4 : Métadonnées ────────────────────────────────────────────
    ws4 = wb.create_sheet("Métadonnées")
    meta = [
        ["Champ", "Valeur"],
        ["Fournisseur", "AgroPrix by 33Lab"],
        ["URL", "https://agroprix.app"],
        ["API", "https://api.agroprix.app/api/v1/"],
        ["Contact", "api@agroprix.app"],
        ["Pays", country_names.get(country, country)],
        ["ISO3", UEMOA_COUNTRIES.get(country, {}).get("iso3", "")],
        ["Filière", commodity.title()],
        ["Période", f"{start_date or 'toutes dates'} → {end_date or _today}"],
        ["Généré le", datetime.now().strftime("%Y-%m-%d %H:%M UTC")],
        ["Généré par", current_user.get("name", "AgroPrix")],
        ["Nb enregistrements prix", len(prices_data)],
        ["Compatible ECOAGRIS", "Oui"],
        ["EUDR-ready", "Oui"],
        ["Source données", "WFP DataBridges (vam-data-bridges 4.0) + terrain"],
        ["Licence", "CC BY 4.0 — Attribution requise : AgroPrix by 33Lab"],
    ]
    for row in meta:
        ws4.append(row)
    style_header(ws4, 1, hdr_fill, hdr_font)
    for row in ws4.iter_rows(min_row=2):
        row[0].font = bold_font
        row[1].font = body_font
        for cell in row:
            cell.border = border
    auto_width(ws4)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"agroprix_{country}_{commodity}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
